"""
Budget Progetto — v3
Novità:
 - Gestione utenti su Google Sheets (aggiunta, reset password, disattivazione)
 - Aggiunta partner non previsti in preventivo con allocazione budget
 - Export consuntivo annuale in Excel e PDF
 - Tabella riepilogativa partner (finanziamento/cofinanziamento + % sul totale)
"""

import streamlit as st
import pandas as pd
from datetime import date, datetime
import gspread
from google.oauth2.service_account import Credentials
import io
import hashlib

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Budget Progetto", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:wght@400;600;700&family=DM+Sans:wght@300;400;500;600&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
h1,h2,h3{font-family:'Fraunces',serif;}
.main-title{font-family:'Fraunces',serif;font-size:2.2rem;color:#1a365d;margin-bottom:0;}
.subtitle{color:#718096;font-size:.9rem;margin-bottom:1.5rem;}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#1a365d 0%,#2a4a7f 100%);}
[data-testid="stSidebar"] *{color:#e2e8f0 !important;}
[data-testid="stSidebar"] label{font-size:.78rem !important;letter-spacing:.06em;text-transform:uppercase;color:#a0aec0 !important;}
[data-testid="metric-container"]{background:white;border:1px solid #e2e8f0;border-radius:12px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);}
.stButton>button{background:linear-gradient(135deg,#2b6cb0,#1a365d);color:white !important;border:none;border-radius:8px;font-weight:600;width:100%;padding:.55rem 1.2rem;}
.stButton>button:hover{opacity:.88;}
.section-header{font-family:'Fraunces',serif;font-size:1.25rem;color:#1a365d;border-left:4px solid #2b6cb0;padding-left:10px;margin:1.8rem 0 1rem 0;}
.badge-fin{background:#ebf8ff;color:#2b6cb0;padding:2px 10px;border-radius:20px;font-size:.78rem;font-weight:600;}
.badge-cofin{background:#f0fff4;color:#276749;padding:2px 10px;border-radius:20px;font-size:.78rem;font-weight:600;}
.alert-success{background:#f0fff4;border-left:4px solid #38a169;padding:12px 16px;border-radius:8px;margin:1rem 0;color:#276749;font-size:.9rem;}
.alert-info{background:#ebf8ff;border-left:4px solid #2b6cb0;padding:12px 16px;border-radius:8px;margin:1rem 0;color:#1a365d;font-size:.9rem;}
.alert-warn{background:#fffbeb;border-left:4px solid #f59e0b;padding:12px 16px;border-radius:8px;margin:1rem 0;color:#92400e;font-size:.9rem;}
.calc-box{background:#f7fafc;border:1px solid #bee3f8;border-radius:10px;padding:14px 18px;margin:.5rem 0 1rem 0;}
.user-card{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:14px 18px;margin:.4rem 0;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# COSTANTI
# ─────────────────────────────────────────────────────────────────────────────
ADMIN_USER = "COORDINATORE"

AREE_DEFAULT = {
    "AREA 1 - Rete Accoglienze": [
        "AZIONE 1.1 - Prima accoglienza","AZIONE 1.2 - Housing Led",
        "AZIONE 1.3 - Una casa per noi","AZIONE 1.4 - Dormitorio invernale"],
    "AREA 2 - Bassa Soglia": [
        "AZIONE 2.1 - Centro diurno","AZIONE 2.2 - Servizi estivi",
        "AZIONE 2.3 - Progetto Includiamo Sul Serio"],
    "AREA 3 - Patti": ["AZIONE 3.1 - Patti"],
    "AREA 4 - Azioni di sistema": [
        "AZIONE 4.1 - Amministrazione di programma",
        "AZIONE 4.2 - Formazione e Comunicazione",
        "AZIONE 4.3 - Monitoraggio e Valutazione",
        "AZIONE 4.4 - Spese generali"],
}

BUDGET_PREVENTIVO_DEFAULT = {
    "AZIONE 1.1 - Prima accoglienza":            (102200.0, 38325.0),
    "AZIONE 1.2 - Housing Led":                  (26400.0,  0.0),
    "AZIONE 1.3 - Una casa per noi":             (13000.0,  0.0),
    "AZIONE 1.4 - Dormitorio invernale":         (5000.0,   145904.18),
    "AZIONE 2.1 - Centro diurno":                (8205.9,   0.0),
    "AZIONE 2.2 - Servizi estivi":               (17525.68, 0.0),
    "AZIONE 2.3 - Progetto Includiamo Sul Serio":(21100.0,  20000.0),
    "AZIONE 3.1 - Patti":                        (27482.8,  0.0),
    "AZIONE 4.1 - Amministrazione di programma": (0.0,      2890.8),
    "AZIONE 4.2 - Formazione e Comunicazione":   (0.0,      3000.0),
    "AZIONE 4.3 - Monitoraggio e Valutazione":   (500.0,    500.0),
    "AZIONE 4.4 - Spese generali":               (6313.44,  0.0),
}

QUADRO_LOGICO_DEFAULT = [
    {"Area":"AREA 1","Azione":"AZIONE 1.1 - Prima accoglienza","Attività":"8 posti distrettuali","Costo":"TOT diaria","RisorseUmane":"educatore, coordinatore, AS","CostoUnitario":35.0,"Quantità":2920,"UdM":"giorni","TotBudget":102200.0,"Partner":"da definirsi","Finanziato":102200.0,"Cofinanziato":0.0},
    {"Area":"AREA 1","Azione":"AZIONE 1.1 - Prima accoglienza","Attività":"3 posti Fond. Madeo","Costo":"TOT diaria","RisorseUmane":"","CostoUnitario":35.0,"Quantità":1095,"UdM":"giorni","TotBudget":38325.0,"Partner":"Fond. Madeo","Finanziato":0.0,"Cofinanziato":38325.0},
    {"Area":"AREA 1","Azione":"AZIONE 1.2 - Housing Led","Attività":"4 appartamenti seconda accoglienza","Costo":"Utenze/affitto","RisorseUmane":"educatore 6h/mese","CostoUnitario":550.0,"Quantità":48,"UdM":"mesi","TotBudget":26400.0,"Partner":"da definirsi","Finanziato":26400.0,"Cofinanziato":0.0},
    {"Area":"AREA 1","Azione":"AZIONE 1.3 - Una casa per noi","Attività":"Rimborso spese appartamento","Costo":"personale + gestione","RisorseUmane":"coordinatore, educatore","CostoUnitario":13000.0,"Quantità":12,"UdM":"mesi","TotBudget":13000.0,"Partner":"Fond. Madeo","Finanziato":13000.0,"Cofinanziato":0.0},
    {"Area":"AREA 1","Azione":"AZIONE 1.4 - Dormitorio invernale","Attività":"Rifugio San Martino","Costo":"personale + utenze","RisorseUmane":"coordinatore, operatori, volontari, AS, psicologo","CostoUnitario":145904.18,"Quantità":6,"UdM":"mesi","TotBudget":145904.18,"Partner":"Fond. Madeo","Finanziato":0.0,"Cofinanziato":145904.18},
    {"Area":"AREA 1","Azione":"AZIONE 1.4 - Dormitorio invernale","Attività":"Potenziamento extra Diocesi","Costo":"personale + gestione","RisorseUmane":"","CostoUnitario":5000.0,"Quantità":12,"UdM":"mesi","TotBudget":5000.0,"Partner":"da definirsi","Finanziato":5000.0,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.1 - Centro diurno","Attività":"Centro diurno","Costo":"personale","RisorseUmane":"educatore Liv 3S","CostoUnitario":24.09,"Quantità":90,"UdM":"h","TotBudget":2168.1,"Partner":"Fond. Madeo","Finanziato":2168.1,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.1 - Centro diurno","Attività":"Centro diurno","Costo":"personale","RisorseUmane":"educatore Liv D3","CostoUnitario":24.69,"Quantità":90,"UdM":"h","TotBudget":2222.1,"Partner":"APG","Finanziato":2222.1,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.1 - Centro diurno","Attività":"Centro diurno","Costo":"personale","RisorseUmane":"educatore Liv D2","CostoUnitario":25.73,"Quantità":90,"UdM":"h","TotBudget":2315.7,"Partner":"Bessimo","Finanziato":2315.7,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.1 - Centro diurno","Attività":"Centro diurno acquisti","Costo":"acquisti","RisorseUmane":"","CostoUnitario":1500.0,"Quantità":1,"UdM":"anno","TotBudget":1500.0,"Partner":"Fond. Madeo","Finanziato":1500.0,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.2 - Servizi estivi","Attività":"Servizio docce","Costo":"personale","RisorseUmane":"educatore Liv 3S","CostoUnitario":24.09,"Quantità":16,"UdM":"gg","TotBudget":770.88,"Partner":"Fond. Madeo","Finanziato":770.88,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.2 - Servizi estivi","Attività":"Servizio docce","Costo":"personale","RisorseUmane":"educatore Liv D3","CostoUnitario":24.69,"Quantità":16,"UdM":"gg","TotBudget":790.08,"Partner":"APG","Finanziato":790.08,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.2 - Servizi estivi","Attività":"Servizio docce","Costo":"personale","RisorseUmane":"educatore Liv D2","CostoUnitario":25.73,"Quantità":16,"UdM":"gg","TotBudget":823.36,"Partner":"Bessimo","Finanziato":823.36,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.2 - Servizi estivi","Attività":"Dormitorio estivo","Costo":"personale","RisorseUmane":"coordinatore, educatore, volontari, AS, psicologo","CostoUnitario":24.09,"Quantità":504,"UdM":"h","TotBudget":12141.36,"Partner":"Fond. Madeo","Finanziato":12141.36,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.2 - Servizi estivi","Attività":"Dormitorio estivo acquisti/utenze","Costo":"acquisti e utenze","RisorseUmane":"","CostoUnitario":2000.0,"Quantità":1,"UdM":"anno","TotBudget":2000.0,"Partner":"Fond. Madeo","Finanziato":2000.0,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.3 - Progetto Includiamo Sul Serio","Attività":"Drop In - utenze/acquisti/affitto","Costo":"utenze, acquisti, affitto","RisorseUmane":"","CostoUnitario":21000.0,"Quantità":12,"UdM":"mesi","TotBudget":21100.0,"Partner":"Bessimo","Finanziato":21100.0,"Cofinanziato":0.0},
    {"Area":"AREA 2","Azione":"AZIONE 2.3 - Progetto Includiamo Sul Serio","Attività":"Drop In - personale","Costo":"personale","RisorseUmane":"4 educatori, 2 medici, 2 infermieri, psicologo, mediatore, ASA, legale, AS","CostoUnitario":20000.0,"Quantità":12,"UdM":"mesi","TotBudget":20000.0,"Partner":"Bessimo","Finanziato":0.0,"Cofinanziato":20000.0},
    {"Area":"AREA 3","Azione":"AZIONE 3.1 - Patti","Attività":"Budget realizzazione patti","Costo":"budget patti","RisorseUmane":"","CostoUnitario":20000.0,"Quantità":12,"UdM":"mesi","TotBudget":20000.0,"Partner":"da definirsi","Finanziato":20000.0,"Cofinanziato":0.0},
    {"Area":"AREA 3","Azione":"AZIONE 3.1 - Patti","Attività":"Personale tutor","Costo":"tutor","RisorseUmane":"educatore, psicologo, AS (6h/mese)","CostoUnitario":26.33,"Quantità":72,"UdM":"ore","TotBudget":5482.8,"Partner":"Fond. Madeo","Finanziato":5482.8,"Cofinanziato":0.0},
    {"Area":"AREA 4","Azione":"AZIONE 4.1 - Amministrazione di programma","Attività":"Costo amministrativo","Costo":"personale","RisorseUmane":"1 amministrativo","CostoUnitario":24.09,"Quantità":120,"UdM":"ore","TotBudget":2890.8,"Partner":"ATS","Finanziato":0.0,"Cofinanziato":2890.8},
    {"Area":"AREA 4","Azione":"AZIONE 4.2 - Formazione e Comunicazione","Attività":"Formazione + comunicazione","Costo":"formazione","RisorseUmane":"","CostoUnitario":3000.0,"Quantità":12,"UdM":"mesi","TotBudget":3000.0,"Partner":"Ufficio di Piano","Finanziato":0.0,"Cofinanziato":3000.0},
    {"Area":"AREA 4","Azione":"AZIONE 4.3 - Monitoraggio e Valutazione","Attività":"Ente esterno monitoraggio","Costo":"monitoraggio","RisorseUmane":"","CostoUnitario":1000.0,"Quantità":12,"UdM":"mesi","TotBudget":1000.0,"Partner":"ATS","Finanziato":500.0,"Cofinanziato":500.0},
    {"Area":"AREA 4","Azione":"AZIONE 4.4 - Spese generali","Attività":"Spese generali 10%","Costo":"spese generali","RisorseUmane":"","CostoUnitario":6313.44,"Quantità":1,"UdM":"anno","TotBudget":6313.44,"Partner":"ATS","Finanziato":6313.44,"Cofinanziato":0.0},
]

TIPI_COSTO = ["Personale - Ore lavorate","Spese - Utenze","Spese - Vitto/Alloggio",
              "Spese - Acquisti","Spese - Affitto","Spese - Altro"]
MESI = ["Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
        "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]
ANNO = 2025

INTESTAZIONE_REND  = ["Timestamp","Partner","Mese","Anno","Area","Azione",
                       "Tipo Costo","Descrizione","Ore","Costo Orario (€)","Importo (€)",
                       "Finanziamento/Cofinanziamento","Note"]
INTESTAZIONE_UTENTI = ["Username","PasswordHash","Ruolo","Attivo","UltimoAccesso"]
SHEET_COSTI  = "CostiOrari"
SHEET_PREV   = "Preventivo"
SHEET_UTENTI = "Utenti"

# ─────────────────────────────────────────────────────────────────────────────
# GOOGLE SHEETS
# ─────────────────────────────────────────────────────────────────────────────
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]

@st.cache_resource
def get_gsheet_client():
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
        raw_key = creds_dict.get("private_key","")
        if raw_key:
            lines   = [l.strip() for l in raw_key.strip().splitlines()]
            header  = next((l for l in lines if "BEGIN" in l),"-----BEGIN PRIVATE KEY-----")
            footer  = next((l for l in lines if "END"   in l),"-----END PRIVATE KEY-----")
            body    = "".join(l for l in lines if "BEGIN" not in l and "END" not in l and l)
            chunked = "\n".join(body[i:i+64] for i in range(0,len(body),64))
            creds_dict["private_key"] = f"{header}\n{chunked}\n{footer}\n"
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"❌ Errore connessione Google Sheets: {e}")
        return None

def get_worksheets(client):
    try:
        sh = client.open_by_key(st.secrets["google_sheet_id"])
        def goc(name, rows, cols, hdr=None):
            try: return sh.worksheet(name)
            except gspread.WorksheetNotFound:
                ws = sh.add_worksheet(title=name, rows=rows, cols=cols)
                if hdr: ws.append_row(hdr)
                return ws
        ws_rend   = goc("Rendicontazione", 5000, 15, INTESTAZIONE_REND)
        ws_costi  = goc(SHEET_COSTI,  500, 5,  ["Partner","FiguraProfessionale","CostoOrario"])
        ws_prev   = goc(SHEET_PREV,   200, 5,  ["Azione","PreventivoFinanziato","PreventivoCofinanziato"])
        ws_utenti = goc(SHEET_UTENTI, 200, 6,  INTESTAZIONE_UTENTI)
        # Seed coordinatore se foglio utenti vuoto
        _init_coordinatore(ws_utenti)
        return ws_rend, ws_costi, ws_prev, ws_utenti
    except Exception as e:
        st.error(f"❌ Errore apertura fogli: {e}"); return None,None,None,None

def _init_coordinatore(ws_utenti):
    records = ws_utenti.get_all_records()
    if not any(r.get("Username") == ADMIN_USER for r in records):
        ws_utenti.append_row([ADMIN_USER, _hash("admin2024"), "admin", "SI",
                               datetime.now().strftime("%d/%m/%Y %H:%M")])

def carica_df(ws, cols):
    try:
        rec = ws.get_all_records()
        return pd.DataFrame(rec) if rec else pd.DataFrame(columns=cols)
    except Exception as e:
        st.error(f"Errore lettura: {e}"); return pd.DataFrame(columns=cols)

def salva_riga(ws, riga):
    try: ws.append_row(riga, value_input_option="USER_ENTERED"); return True
    except Exception as e: st.error(f"Errore: {e}"); return False

# ─────────────────────────────────────────────────────────────────────────────
# AUTH HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _hash(pwd): return hashlib.sha256(pwd.encode()).hexdigest()

def carica_utenti(ws_utenti):
    return carica_df(ws_utenti, INTESTAZIONE_UTENTI)

def verifica_login(ws_utenti, username, password):
    df = carica_utenti(ws_utenti)
    if df.empty: return False, None
    row = df[(df["Username"] == username) & (df["Attivo"] == "SI")]
    if row.empty: return False, None
    if row.iloc[0]["PasswordHash"] == _hash(password):
        return True, row.iloc[0]["Ruolo"]
    return False, None

def aggiorna_ultimo_accesso(ws_utenti, username):
    records = ws_utenti.get_all_records()
    for i, r in enumerate(records, start=2):
        if r.get("Username") == username:
            ws_utenti.update_cell(i, 5, datetime.now().strftime("%d/%m/%Y %H:%M"))
            break

def reset_password(ws_utenti, username, nuova_pwd):
    records = ws_utenti.get_all_records()
    for i, r in enumerate(records, start=2):
        if r.get("Username") == username:
            ws_utenti.update_cell(i, 2, _hash(nuova_pwd))
            return True
    return False

def imposta_attivo(ws_utenti, username, attivo: bool):
    records = ws_utenti.get_all_records()
    for i, r in enumerate(records, start=2):
        if r.get("Username") == username:
            ws_utenti.update_cell(i, 4, "SI" if attivo else "NO")
            return True
    return False

def aggiungi_utente(ws_utenti, username, password, ruolo="partner"):
    df = carica_utenti(ws_utenti)
    if not df.empty and username in df["Username"].values:
        return False, "Username già esistente."
    ws_utenti.append_row([username, _hash(password), ruolo, "SI",
                           datetime.now().strftime("%d/%m/%Y %H:%M")])
    return True, "OK"

# ─────────────────────────────────────────────────────────────────────────────
# PREVENTIVO / COSTI ORARI HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def carica_preventivo(ws_prev):
    df = carica_df(ws_prev, ["Azione","PreventivoFinanziato","PreventivoCofinanziato"])
    if df.empty: return dict(BUDGET_PREVENTIVO_DEFAULT)
    r = {}
    for _, row in df.iterrows():
        try: r[str(row["Azione"]).strip()] = (float(row["PreventivoFinanziato"]), float(row["PreventivoCofinanziato"]))
        except: pass
    return r if r else dict(BUDGET_PREVENTIVO_DEFAULT)

def carica_costi_orari(ws_costi):
    df = carica_df(ws_costi, ["Partner","FiguraProfessionale","CostoOrario"])
    r = {}
    for _, row in df.iterrows():
        try: r[(str(row["Partner"]).strip(), str(row["FiguraProfessionale"]).strip())] = float(row["CostoOrario"])
        except: pass
    return r

def figure_per_partner(co, partner):
    return [f for (p,f) in co.keys() if p == partner]

def aree_da_preventivo(budget):
    aree = {}
    for az in budget:
        try:
            num = az.split("AZIONE")[1].strip().split(".")[0].strip()
            label = next((k for k in AREE_DEFAULT if k.startswith(f"AREA {num}")), f"AREA {num}")
        except: label = "Altre azioni"
        aree.setdefault(label, [])
        if az not in aree[label]: aree[label].append(az)
    return aree if aree else AREE_DEFAULT

def utenti_partner(ws_utenti):
    df = carica_utenti(ws_utenti)
    if df.empty: return []
    return df[df["Ruolo"] == "partner"]["Username"].tolist()

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def genera_excel_consuntivo(df_rend, budget_prev, anno):
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Foglio 1: Consuntivo per azione ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Consuntivo per Azione"

    BLUE  = "1a365d"; LBLUE = "2b6cb0"; GREEN = "276749"
    HBLU  = PatternFill("solid", fgColor=BLUE)
    HLBLU = PatternFill("solid", fgColor="EBF8FF")
    HGRN  = PatternFill("solid", fgColor="F0FFF4")
    HTOT  = PatternFill("solid", fgColor="E2E8F0")
    thin  = Side(style="thin", color="CBD5E0")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, val, fill=HBLU, bold=True, color="FFFFFF", wrap=True):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = Font(bold=bold, color=color, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = brd; return c

    def data_cell(ws, row, col, val, num_fmt=None, fill=None, bold=False, color="2D3748", align="left"):
        c = ws.cell(row=row, column=col, value=val)
        if fill: c.fill = fill
        c.font = Font(bold=bold, color=color, size=9)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = brd
        if num_fmt: c.number_format = num_fmt
        return c

    # Titolo
    ws1.merge_cells("A1:N1")
    tc = ws1["A1"]; tc.value = f"QUADRO LOGICO CONSUNTIVO {anno}"
    tc.font = Font(bold=True, size=14, color=BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill = PatternFill("solid", fgColor="EBF4FF")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:N2")
    sc = ws1["A2"]; sc.value = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sc.font = Font(italic=True, size=8, color="718096")
    sc.alignment = Alignment(horizontal="center")

    hdrs = ["Area","Azione","Partner","Tipo Costo","Descrizione",
            "Ore Tot.","Prev. Fin. (€)","Prev. Cofin. (€)","Prev. Tot. (€)",
            "Cons. Fin. (€)","Cons. Cofin. (€)","Cons. Tot. (€)",
            "Scost. Fin. (€)","Scost. Cofin. (€)"]
    for ci, h in enumerate(hdrs, 1):
        hdr_cell(ws1, 3, ci, h)
    ws1.row_dimensions[3].height = 32

    # Dati per azione/partner
    df_r = df_rend.copy()
    df_r["Importo (€)"] = pd.to_numeric(df_r["Importo (€)"], errors="coerce").fillna(0)

    row_i = 4
    for azione, (pf, pc) in budget_prev.items():
        df_az = df_r[df_r["Azione"] == azione]
        partners_az = df_az["Partner"].unique().tolist() if not df_az.empty else ["—"]
        if not partners_az: partners_az = ["—"]

        first = True
        for partner in partners_az:
            df_p = df_az[df_az["Partner"] == partner] if partner != "—" else pd.DataFrame()
            for tipo in (df_p["Tipo Costo"].unique().tolist() if not df_p.empty else ["—"]):
                df_t  = df_p[df_p["Tipo Costo"] == tipo] if not df_p.empty else pd.DataFrame()
                desc  = "; ".join(df_t["Descrizione"].dropna().unique()[:3]) if not df_t.empty else ""
                ore   = df_t["Ore"].apply(pd.to_numeric, errors="coerce").sum() if not df_t.empty else 0
                cf    = df_t[df_t["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum() if not df_t.empty else 0
                cc    = df_t[df_t["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum() if not df_t.empty else 0

                area_label = next((k for k in AREE_DEFAULT if azione in AREE_DEFAULT.get(k,[])), "")
                fill_row = HLBLU if row_i % 2 == 0 else None

                data_cell(ws1, row_i, 1,  area_label, fill=fill_row)
                data_cell(ws1, row_i, 2,  azione,      fill=fill_row)
                data_cell(ws1, row_i, 3,  partner,     fill=fill_row)
                data_cell(ws1, row_i, 4,  tipo,        fill=fill_row)
                data_cell(ws1, row_i, 5,  desc,        fill=fill_row)
                data_cell(ws1, row_i, 6,  ore,         num_fmt="#,##0.0", fill=fill_row, align="right")
                if first:
                    data_cell(ws1, row_i, 7,  pf,  num_fmt='#,##0.00 "€"', fill=fill_row, align="right")
                    data_cell(ws1, row_i, 8,  pc,  num_fmt='#,##0.00 "€"', fill=fill_row, align="right")
                    data_cell(ws1, row_i, 9,  pf+pc, num_fmt='#,##0.00 "€"', fill=fill_row, align="right", bold=True)
                else:
                    for ci in [7,8,9]: data_cell(ws1, row_i, ci, "", fill=fill_row)
                data_cell(ws1, row_i, 10, cf,      num_fmt='#,##0.00 "€"', fill=fill_row, align="right",
                          color="276749" if cf>0 else "2D3748")
                data_cell(ws1, row_i, 11, cc,      num_fmt='#,##0.00 "€"', fill=fill_row, align="right",
                          color="2B6CB0" if cc>0 else "2D3748")
                data_cell(ws1, row_i, 12, cf+cc,   num_fmt='#,##0.00 "€"', fill=fill_row, align="right", bold=True)
                data_cell(ws1, row_i, 13, cf-pf if first else "",
                          num_fmt='#,##0.00 "€"', fill=fill_row, align="right",
                          color="E53E3E" if isinstance(cf-pf,float) and cf-pf<0 else "276749")
                data_cell(ws1, row_i, 14, cc-pc if first else "",
                          num_fmt='#,##0.00 "€"', fill=fill_row, align="right",
                          color="E53E3E" if isinstance(cc-pc,float) and cc-pc<0 else "276749")
                row_i += 1
                first  = False

    # Totale
    tot_pf = sum(v[0] for v in budget_prev.values())
    tot_pc = sum(v[1] for v in budget_prev.values())
    tot_cf = df_r[df_r["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
    tot_cc = df_r[df_r["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum()
    for ci, (val, fmt) in enumerate([
        ("",""), ("TOTALE GENERALE",""), ("",""), ("",""), ("",""),
        (df_r["Ore"].apply(pd.to_numeric,errors="coerce").sum(), "#,##0.0"),
        (tot_pf,'#,##0.00 "€"'), (tot_pc,'#,##0.00 "€"'), (tot_pf+tot_pc,'#,##0.00 "€"'),
        (tot_cf,'#,##0.00 "€"'), (tot_cc,'#,##0.00 "€"'), (tot_cf+tot_cc,'#,##0.00 "€"'),
        (tot_cf-tot_pf,'#,##0.00 "€"'), (tot_cc-tot_pc,'#,##0.00 "€"')
    ], 1):
        c = ws1.cell(row=row_i, column=ci, value=val)
        c.fill = HTOT; c.font = Font(bold=True, size=9, color=BLUE)
        c.alignment = Alignment(horizontal="right" if ci>2 else "left", vertical="center")
        c.border = brd
        if fmt: c.number_format = fmt

    # Larghezze colonne
    for ci, w in enumerate([18,32,18,20,30,8,14,14,14,14,14,14,14,14], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A4"

    # ── Foglio 2: Riepilogo per partner ───────────────────────────────────
    ws2 = wb.create_sheet("Riepilogo Partner")
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]; t2.value = f"RIEPILOGO PER PARTNER — {anno}"
    t2.font = Font(bold=True, size=13, color=BLUE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t2.fill = PatternFill("solid", fgColor="EBF4FF")
    ws2.row_dimensions[1].height = 28

    h2 = ["Partner","Finanziamento (€)","% su Tot. Fin.","Cofinanziamento (€)",
          "% su Tot. Cofin.","Totale (€)","% su Tot. Progetto"]
    for ci, h in enumerate(h2, 1):
        hdr_cell(ws2, 2, ci, h)
    ws2.row_dimensions[2].height = 28

    df_sum = df_r.groupby(["Partner","Finanziamento/Cofinanziamento"])["Importo (€)"].sum().unstack(fill_value=0).reset_index()
    if "Finanziamento"  not in df_sum.columns: df_sum["Finanziamento"]  = 0
    if "Cofinanziamento" not in df_sum.columns: df_sum["Cofinanziamento"] = 0
    df_sum["Totale"] = df_sum["Finanziamento"] + df_sum["Cofinanziamento"]
    df_sum = df_sum.sort_values("Totale", ascending=False)

    grand_fin   = df_sum["Finanziamento"].sum()
    grand_cofin = df_sum["Cofinanziamento"].sum()
    grand_tot   = grand_fin + grand_cofin

    ri = 3
    for _, row in df_sum.iterrows():
        fill_r = PatternFill("solid", fgColor="F7FAFC") if ri % 2 == 0 else None
        pct_f  = row["Finanziamento"]  / grand_fin   * 100 if grand_fin   > 0 else 0
        pct_c  = row["Cofinanziamento"]/ grand_cofin * 100 if grand_cofin > 0 else 0
        pct_t  = row["Totale"]         / grand_tot   * 100 if grand_tot   > 0 else 0
        data_cell(ws2, ri, 1, row["Partner"], fill=fill_r, bold=True)
        data_cell(ws2, ri, 2, row["Finanziamento"],  '#,##0.00 "€"', fill=fill_r, align="right", color="276749")
        data_cell(ws2, ri, 3, pct_f/100,            "0.0%",          fill=fill_r, align="right")
        data_cell(ws2, ri, 4, row["Cofinanziamento"],'#,##0.00 "€"', fill=fill_r, align="right", color=LBLUE)
        data_cell(ws2, ri, 5, pct_c/100,            "0.0%",          fill=fill_r, align="right")
        data_cell(ws2, ri, 6, row["Totale"],         '#,##0.00 "€"', fill=fill_r, align="right", bold=True)
        data_cell(ws2, ri, 7, pct_t/100,            "0.0%",          fill=fill_r, align="right")
        ri += 1

    # Riga totale
    for ci, (val, fmt) in enumerate([
        ("TOTALE",""),
        (grand_fin,   '#,##0.00 "€"'), (1.0, "0.0%"),
        (grand_cofin, '#,##0.00 "€"'), (1.0, "0.0%"),
        (grand_tot,   '#,##0.00 "€"'), (1.0, "0.0%"),
    ], 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.fill = HTOT; c.font = Font(bold=True, size=9, color=BLUE)
        c.alignment = Alignment(horizontal="right" if ci>1 else "left", vertical="center")
        c.border = brd
        if fmt: c.number_format = fmt

    for ci, w in enumerate([22,18,12,18,12,16,16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A3"

    # ── Foglio 3: Dati grezzi ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Dati Grezzi")
    for ci, h in enumerate(df_rend.columns, 1):
        hdr_cell(ws3, 1, ci, h)
    for ri2, row in enumerate(df_rend.values, 2):
        for ci, val in enumerate(row, 1):
            ws3.cell(row=ri2, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PDF
# ─────────────────────────────────────────────────────────────────────────────
def genera_pdf_consuntivo(df_rend, budget_prev, anno):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    C_BLUE  = colors.HexColor("#1a365d")
    C_LBLUE = colors.HexColor("#2b6cb0")
    C_GREEN = colors.HexColor("#276749")
    C_RED   = colors.HexColor("#c53030")
    C_LGRAY = colors.HexColor("#f7fafc")
    C_EGRAY = colors.HexColor("#e2e8f0")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=C_BLUE, fontSize=16, spaceAfter=4)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                  textColor=colors.HexColor("#718096"), fontSize=9, spaceAfter=12)
    h2_style    = ParagraphStyle("h2", parent=styles["Heading2"],
                                  textColor=C_BLUE, fontSize=12, spaceBefore=16, spaceAfter=6)

    df_r = df_rend.copy()
    df_r["Importo (€)"] = pd.to_numeric(df_r["Importo (€)"], errors="coerce").fillna(0)

    story = []
    story.append(Paragraph(f"Quadro Logico Consuntivo {anno}", title_style))
    story.append(Paragraph(f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))

    # ── Sezione 1: Riepilogo per partner ─────────────────────────────────
    story.append(Paragraph("Riepilogo per Partner", h2_style))

    df_sum = df_r.groupby(["Partner","Finanziamento/Cofinanziamento"])["Importo (€)"].sum().unstack(fill_value=0).reset_index()
    if "Finanziamento"   not in df_sum.columns: df_sum["Finanziamento"]   = 0
    if "Cofinanziamento" not in df_sum.columns: df_sum["Cofinanziamento"] = 0
    df_sum["Totale"] = df_sum["Finanziamento"] + df_sum["Cofinanziamento"]
    df_sum = df_sum.sort_values("Totale", ascending=False)
    grand_fin   = df_sum["Finanziamento"].sum()
    grand_cofin = df_sum["Cofinanziamento"].sum()
    grand_tot   = grand_fin + grand_cofin

    t_data = [["Partner","Finanziamento","% Fin.","Cofinanziamento","% Cofin.","Totale","% Tot."]]
    for _, row in df_sum.iterrows():
        pf = row["Finanziamento"]/grand_fin*100   if grand_fin>0   else 0
        pc = row["Cofinanziamento"]/grand_cofin*100 if grand_cofin>0 else 0
        pt = row["Totale"]/grand_tot*100           if grand_tot>0   else 0
        t_data.append([
            row["Partner"],
            f"€ {row['Finanziamento']:,.2f}",  f"{pf:.1f}%",
            f"€ {row['Cofinanziamento']:,.2f}", f"{pc:.1f}%",
            f"€ {row['Totale']:,.2f}",          f"{pt:.1f}%",
        ])
    t_data.append([
        "TOTALE",
        f"€ {grand_fin:,.2f}", "100%",
        f"€ {grand_cofin:,.2f}", "100%",
        f"€ {grand_tot:,.2f}", "100%",
    ])

    col_w = [5*cm, 3.2*cm, 1.8*cm, 3.2*cm, 1.8*cm, 3.2*cm, 1.8*cm]
    t_partner = Table(t_data, colWidths=col_w)
    t_partner.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_BLUE),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ALIGN",      (1,0), (-1,-1), "RIGHT"),
        ("ALIGN",      (0,0), (0,-1), "LEFT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [C_LGRAY, colors.white]),
        ("BACKGROUND", (0,-1), (-1,-1), C_EGRAY),
        ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E0")),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    story.append(t_partner)
    story.append(PageBreak())

    # ── Sezione 2: Consuntivo per azione ─────────────────────────────────
    story.append(Paragraph("Consuntivo per Azione", h2_style))

    a_data = [["Azione","Partner","Prev. Fin.","Prev. Cofin.",
               "Cons. Fin.","Cons. Cofin.","Cons. Tot.","Scost. Fin."]]
    for azione, (pf, pc) in budget_prev.items():
        df_az = df_r[df_r["Azione"] == azione]
        cf = df_az[df_az["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
        cc = df_az[df_az["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum()
        partners_list = df_az["Partner"].unique().tolist() if not df_az.empty else ["—"]
        a_data.append([
            azione[:40],
            ", ".join(partners_list)[:30],
            f"€ {pf:,.0f}",  f"€ {pc:,.0f}",
            f"€ {cf:,.0f}",  f"€ {cc:,.0f}",
            f"€ {cf+cc:,.0f}",
            f"€ {cf-pf:+,.0f}",
        ])

    tot_pf = sum(v[0] for v in budget_prev.values())
    tot_pc = sum(v[1] for v in budget_prev.values())
    tot_cf = df_r[df_r["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
    tot_cc = df_r[df_r["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum()
    a_data.append([
        "TOTALE","",
        f"€ {tot_pf:,.0f}", f"€ {tot_pc:,.0f}",
        f"€ {tot_cf:,.0f}", f"€ {tot_cc:,.0f}",
        f"€ {tot_cf+tot_cc:,.0f}",
        f"€ {tot_cf-tot_pf:+,.0f}",
    ])

    cw2 = [6.5*cm, 3.8*cm, 2.6*cm, 2.6*cm, 2.6*cm, 2.6*cm, 2.6*cm, 2.6*cm]
    t_azioni = Table(a_data, colWidths=cw2)

    style_azioni = [
        ("BACKGROUND", (0,0), (-1,0), C_LBLUE),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 7.5),
        ("ALIGN",      (2,0), (-1,-1), "RIGHT"),
        ("ALIGN",      (0,0), (1,-1), "LEFT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [C_LGRAY, colors.white]),
        ("BACKGROUND", (0,-1), (-1,-1), C_EGRAY),
        ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E0")),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]
    # Colora scostamenti negativi in rosso
    for ri2, azione in enumerate(list(budget_prev.keys()), 1):
        df_az = df_r[df_r["Azione"] == azione]
        cf = df_az[df_az["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
        pf = budget_prev[azione][0]
        if cf - pf < 0:
            style_azioni.append(("TEXTCOLOR", (7, ri2), (7, ri2), C_RED))

    t_azioni.setStyle(TableStyle(style_azioni))
    story.append(t_azioni)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for k, v in [("logged_in",False),("partner",None),("ruolo",None)]:
    if k not in st.session_state: st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def mostra_login(ws_utenti):
    _, col_m, _ = st.columns([1, 1.2, 1])
    with col_m:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown('<div class="main-title" style="text-align:center">📊 Budget Progetto</div>', unsafe_allow_html=True)
        st.markdown('<div class="subtitle" style="text-align:center">Monitoraggio spese e ore per partner</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("#### Accedi al tuo spazio")
            username = st.text_input("Nome utente")
            password = st.text_input("Password", type="password")
            if st.button("Accedi →"):
                ok, ruolo = verifica_login(ws_utenti, username.strip(), password)
                if ok:
                    st.session_state.logged_in = True
                    st.session_state.partner   = username.strip()
                    st.session_state.ruolo     = ruolo
                    aggiorna_ultimo_accesso(ws_utenti, username.strip())
                    st.rerun()
                else:
                    st.error("Credenziali errate o account disattivato.")

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
def mostra_sidebar():
    with st.sidebar:
        st.markdown(f"### 👤 {st.session_state.partner}")
        ruolo = st.session_state.ruolo
        st.markdown(f"<small style='color:#a0aec0'>Ruolo: {ruolo}</small>", unsafe_allow_html=True)
        st.markdown("---")
        if ruolo == "admin":
            voci = ["📥 Inserimento","📊 Cruscotto","📋 Quadro logico",
                    "⚙️ Gestione preventivo","👥 Gestione utenti","📤 Export consuntivo"]
        else:
            voci = ["📥 Inserimento","📋 Quadro logico"]
        pagina = st.radio("Navigazione", voci, label_visibility="collapsed")
        st.markdown("---")
        if st.button("🔓 Esci"):
            st.session_state.logged_in = False
            st.session_state.partner   = None
            st.session_state.ruolo     = None
            st.rerun()
    return pagina

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA INSERIMENTO
# ─────────────────────────────────────────────────────────────────────────────
def pagina_inserimento(ws_rend, ws_costi, ws_prev, partner):
    st.markdown('<div class="main-title">📥 Inserimento mensile</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="subtitle">Ente: <strong>{partner}</strong> — Anno {ANNO}</div>', unsafe_allow_html=True)

    budget = carica_preventivo(ws_prev)
    aree   = aree_da_preventivo(budget)
    co     = carica_costi_orari(ws_costi)

    col1, col2 = st.columns(2)
    with col1: mese     = st.selectbox("Mese", MESI, index=date.today().month-1)
    with col2: area_sel = st.selectbox("Area", list(aree.keys()))
    azione_sel = st.selectbox("Azione", aree[area_sel])

    st.markdown('<div class="section-header">Voce di spesa / ore</div>', unsafe_allow_html=True)
    col3, col4 = st.columns(2)
    with col3: tipo_costo = st.selectbox("Tipo di costo", TIPI_COSTO)
    with col4: fin_cofin  = st.radio("Imputazione", ["Finanziamento","Cofinanziamento"], horizontal=True)

    descrizione = st.text_input("Descrizione voce", placeholder="es. Mario Rossi - Educatore")

    ore = costo_orario_usato = 0.0
    if "Personale" in tipo_costo:
        figure = figure_per_partner(co, partner)
        c5, c6, c7 = st.columns(3)
        with c5: ore = st.number_input("Ore lavorate nel mese", min_value=0.0, step=0.5, format="%.1f")
        with c6:
            if figure:
                scelta = st.selectbox("Figura professionale", ["— seleziona —"]+figure)
                costo_orario_usato = co.get((partner, scelta), 0.0) if scelta != "— seleziona —" else st.number_input("Costo orario (€/h)", min_value=0.0, step=0.01, format="%.2f", key="co_manual")
                if scelta != "— seleziona —": st.caption(f"Tariffa: € {costo_orario_usato:.2f}/h")
            else:
                costo_orario_usato = st.number_input("Costo orario (€/h)", min_value=0.0, step=0.01, format="%.2f",
                    help="Configura le tariffe in '⚙️ Gestione preventivo'")
        importo_calc = round(ore * costo_orario_usato, 2)
        with c7:
            st.markdown(f"""<div class='calc-box'>
              <small style='color:#718096;text-transform:uppercase;font-size:.75rem'>Importo calcolato</small><br>
              <span style='font-size:1.6rem;font-weight:700;color:#1a365d'>€ {importo_calc:,.2f}</span><br>
              <small style='color:#a0aec0'>{ore:.1f} h × € {costo_orario_usato:.2f}/h</small>
            </div>""", unsafe_allow_html=True)
        importo = st.number_input("Importo (€) — modificabile", min_value=0.0,
                                   value=float(importo_calc), step=0.01, format="%.2f")
    else:
        importo = st.number_input("Importo (€)", min_value=0.0, step=0.01, format="%.2f")

    note = st.text_area("Note", placeholder="Riferimenti, giustificativi...", height=70)

    pf, pc = budget.get(azione_sel, (0,0))
    with st.expander("📋 Preventivo per questa azione"):
        c1, c2, c3 = st.columns(3)
        c1.metric("Prev. finanziato",   f"€ {pf:,.2f}")
        c2.metric("Prev. cofinanziato", f"€ {pc:,.2f}")
        c3.metric("Totale azione",      f"€ {pf+pc:,.2f}")

    st.markdown("---")
    if st.button("💾 Salva voce"):
        if not descrizione.strip():    st.warning("Inserisci una descrizione."); return
        if importo == 0 and ore == 0:  st.warning("Inserisci importo o ore."); return
        riga = [datetime.now().strftime("%d/%m/%Y %H:%M"), partner, mese, ANNO,
                area_sel, azione_sel, tipo_costo, descrizione.strip(),
                ore, costo_orario_usato, importo, fin_cofin, note.strip()]
        if salva_riga(ws_rend, riga):
            st.markdown('<div class="alert-success">✅ Voce salvata!</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header">Le tue voci inserite</div>', unsafe_allow_html=True)
    df = carica_df(ws_rend, INTESTAZIONE_REND)
    if not df.empty:
        df_p = df[df["Partner"] == partner].copy()
        if not df_p.empty:
            df_p["Importo (€)"] = pd.to_numeric(df_p["Importo (€)"], errors="coerce").map(
                lambda x: f"€ {x:,.2f}" if pd.notna(x) else "")
            st.dataframe(df_p[["Mese","Azione","Tipo Costo","Descrizione","Ore",
                                "Costo Orario (€)","Importo (€)","Finanziamento/Cofinanziamento"]],
                         use_container_width=True, hide_index=True)
        else: st.info("Nessuna voce ancora inserita.")
    else: st.info("Nessuna voce ancora inserita.")

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA QUADRO LOGICO
# ─────────────────────────────────────────────────────────────────────────────
def pagina_quadro_logico(partner, ruolo):
    is_admin = (ruolo == "admin")
    st.markdown('<div class="main-title">📋 Quadro Logico</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="subtitle">{"Vista completa" if is_admin else f"Voci di competenza: <strong>{partner}</strong>"} — sola lettura</div>', unsafe_allow_html=True)

    df_ql = pd.DataFrame(QUADRO_LOGICO_DEFAULT)
    if not is_admin:
        df_ql = df_ql[df_ql["Partner"].str.strip() == partner]
        if df_ql.empty:
            st.info("Nessuna voce del quadro logico associata a questo account. Contatta il coordinatore.")
            return

    tot_fin   = df_ql["Finanziato"].sum()
    tot_cofin = df_ql["Cofinanziato"].sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("💶 Finanziato",    f"€ {tot_fin:,.2f}")
    c2.metric("🤝 Cofinanziato",  f"€ {tot_cofin:,.2f}")
    c3.metric("📊 Totale",        f"€ {tot_fin+tot_cofin:,.2f}")

    for area in df_ql["Area"].unique():
        st.markdown(f'<div class="section-header">{area}</div>', unsafe_allow_html=True)
        df_a = df_ql[df_ql["Area"] == area].copy()
        df_a["CostoUnitario"] = df_a["CostoUnitario"].map(lambda x: f"€ {x:,.2f}")
        df_a["TotBudget"]    = df_a["TotBudget"].map(lambda x: f"€ {x:,.2f}")
        df_a["Finanziato"]   = df_a["Finanziato"].map(lambda x: f"€ {x:,.2f}")
        df_a["Cofinanziato"] = df_a["Cofinanziato"].map(lambda x: f"€ {x:,.2f}")
        cols = ["Azione","Attività","Costo","RisorseUmane","CostoUnitario",
                "Quantità","UdM","TotBudget","Finanziato","Cofinanziato"]
        if is_admin: cols.insert(8,"Partner")
        st.dataframe(df_a[cols], use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA GESTIONE PREVENTIVO
# ─────────────────────────────────────────────────────────────────────────────
def pagina_gestione_preventivo(ws_costi, ws_prev, ws_utenti):
    st.markdown('<div class="main-title">⚙️ Gestione preventivo</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Personalizza budget, costi orari e aggiungi partner</div>', unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["📊 Carica da Excel","✏️ Modifica manuale",
                                       "💼 Costi orari","➕ Nuovo partner"])

    with tab1:
        st.markdown('<div class="section-header">Importa preventivo da Excel</div>', unsafe_allow_html=True)
        st.markdown("""<div class='alert-info'>Colonne richieste: <strong>Azione</strong> | <strong>PreventivoFinanziato</strong> | <strong>PreventivoCofinanziato</strong></div>""", unsafe_allow_html=True)
        up = st.file_uploader("Carica .xlsx", type=["xlsx","xls"])
        if up:
            try:
                df_xl = pd.read_excel(up)
                st.dataframe(df_xl.head(6), use_container_width=True, hide_index=True)
                c_az  = st.selectbox("Colonna Azione",                 df_xl.columns.tolist(), key="xa")
                c_fin = st.selectbox("Colonna Preventivo Finanziato",  df_xl.columns.tolist(), key="xf")
                c_cof = st.selectbox("Colonna Preventivo Cofinanziato",df_xl.columns.tolist(), key="xc")
                if st.button("💾 Importa"):
                    ws_prev.clear(); ws_prev.append_row(["Azione","PreventivoFinanziato","PreventivoCofinanziato"])
                    n = 0
                    for _, row in df_xl.iterrows():
                        az = str(row[c_az]).strip()
                        if not az or az.lower()=="nan": continue
                        try: fin=float(str(row[c_fin]).replace("€","").replace(",",".") or 0)
                        except: fin=0.0
                        try: cof=float(str(row[c_cof]).replace("€","").replace(",",".") or 0)
                        except: cof=0.0
                        ws_prev.append_row([az, fin, cof]); n+=1
                    st.success(f"✅ {n} azioni importate!")
            except Exception as e: st.error(f"Errore: {e}")

    with tab2:
        st.markdown('<div class="section-header">Preventivo attuale</div>', unsafe_allow_html=True)
        budget = carica_preventivo(ws_prev)
        df_p = pd.DataFrame([{"Azione":k,"Finanziato (€)":v[0],"Cofinanziato (€)":v[1],"Tot (€)":v[0]+v[1]}
                               for k,v in budget.items()])
        st.dataframe(df_p, use_container_width=True, hide_index=True)
        st.markdown('<div class="section-header">Aggiungi / modifica azione</div>', unsafe_allow_html=True)
        az_n  = st.text_input("Nome azione")
        cc1, cc2 = st.columns(2)
        with cc1: fin_n  = st.number_input("Prev. finanziato (€)",   min_value=0.0, step=100.0, format="%.2f", key="mn_f")
        with cc2: cofin_n= st.number_input("Prev. cofinanziato (€)", min_value=0.0, step=100.0, format="%.2f", key="mn_c")
        if st.button("💾 Salva azione"):
            if not az_n.strip(): st.warning("Inserisci il nome dell'azione.")
            else:
                recs = ws_prev.get_all_records(); found=False
                for i,r in enumerate(recs, start=2):
                    if str(r.get("Azione","")).strip()==az_n.strip():
                        ws_prev.update(f"A{i}:C{i}", [[az_n.strip(), fin_n, cofin_n]]); found=True; break
                if not found: ws_prev.append_row([az_n.strip(), fin_n, cofin_n])
                st.success(f"✅ '{az_n}' salvata!"); st.rerun()

    with tab3:
        st.markdown('<div class="section-header">Tariffe orarie configurate</div>', unsafe_allow_html=True)
        df_co = carica_df(ws_costi, ["Partner","FiguraProfessionale","CostoOrario"])
        if not df_co.empty:
            df_co_disp = df_co.copy()
            df_co_disp["CostoOrario"] = pd.to_numeric(df_co_disp["CostoOrario"], errors="coerce").map(
                lambda x: f"€ {x:.2f}/h" if pd.notna(x) else "")
            st.dataframe(df_co_disp, use_container_width=True, hide_index=True)
        else: st.info("Nessuna tariffa configurata.")
        st.markdown('<div class="section-header">Aggiungi tariffa</div>', unsafe_allow_html=True)
        all_partners = utenti_partner(ws_utenti)
        p_co  = st.selectbox("Partner", all_partners if all_partners else ["— nessun partner —"])
        fig_co = st.text_input("Figura professionale (es. Educatore Liv.3S)")
        tar_co = st.number_input("Costo orario (€/h)", min_value=0.0, step=0.01, format="%.2f")
        if st.button("💾 Salva tariffa"):
            if not fig_co.strip(): st.warning("Inserisci la figura professionale.")
            elif tar_co<=0: st.warning("Costo orario > 0.")
            else:
                recs=ws_costi.get_all_records(); found=False
                for i,r in enumerate(recs, start=2):
                    if str(r.get("Partner","")).strip()==p_co and str(r.get("FiguraProfessionale","")).strip()==fig_co.strip():
                        ws_costi.update(f"A{i}:C{i}",[[p_co,fig_co.strip(),tar_co]]); found=True; break
                if not found: ws_costi.append_row([p_co, fig_co.strip(), tar_co])
                st.success(f"✅ {p_co} — {fig_co} → € {tar_co:.2f}/h"); st.rerun()
        if not df_co.empty:
            st.markdown('<div class="section-header">Elimina tariffa</div>', unsafe_allow_html=True)
            opz = [f"{r['Partner']} — {r['FiguraProfessionale']}" for _,r in df_co.iterrows()]
            da_el = st.selectbox("Seleziona", opz)
            if st.button("🗑️ Elimina"):
                ws_costi.delete_rows(opz.index(da_el)+2); st.success("Eliminata."); st.rerun()

    with tab4:
        st.markdown('<div class="section-header">Aggiungi partner non previsto</div>', unsafe_allow_html=True)
        st.markdown("""<div class='alert-info'>Qui puoi aggiungere un partner nuovo: verrà creato un account utente
        e potrai allocargli subito un budget su un'azione specifica.</div>""", unsafe_allow_html=True)

        nuovo_nome = st.text_input("Nome ente / partner")
        nuova_pwd  = st.text_input("Password iniziale", type="password")
        nuova_pwd2 = st.text_input("Conferma password", type="password")

        st.markdown("**Allocazione budget (opzionale)**")
        budget = carica_preventivo(ws_prev)
        azioni_list = list(budget.keys())
        az_alloc = st.selectbox("Azione su cui allocare budget", ["— nessuna —"]+azioni_list)
        ca1, ca2 = st.columns(2)
        with ca1: alloc_fin  = st.number_input("Budget finanziato (€)",   min_value=0.0, step=100.0, format="%.2f", key="af")
        with ca2: alloc_cofin= st.number_input("Budget cofinanziato (€)", min_value=0.0, step=100.0, format="%.2f", key="ac")

        if st.button("➕ Aggiungi partner"):
            if not nuovo_nome.strip(): st.warning("Inserisci il nome del partner."); st.stop()
            if not nuova_pwd:          st.warning("Inserisci una password."); st.stop()
            if nuova_pwd != nuova_pwd2: st.error("Le password non coincidono."); st.stop()
            ok, msg = aggiungi_utente(ws_utenti, nuovo_nome.strip(), nuova_pwd)
            if not ok: st.error(msg); st.stop()
            # Allocazione budget
            if az_alloc != "— nessuna —" and (alloc_fin > 0 or alloc_cofin > 0):
                pf_ex, pc_ex = budget.get(az_alloc, (0.0, 0.0))
                recs = ws_prev.get_all_records(); found=False
                for i,r in enumerate(recs, start=2):
                    if str(r.get("Azione","")).strip() == az_alloc:
                        ws_prev.update(f"A{i}:C{i}", [[az_alloc, pf_ex+alloc_fin, pc_ex+alloc_cofin]])
                        found=True; break
                if not found:
                    ws_prev.append_row([az_alloc, alloc_fin, alloc_cofin])
            st.success(f"✅ Partner '{nuovo_nome}' aggiunto! Credenziali: username = {nuovo_nome.strip()}")
            st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA GESTIONE UTENTI
# ─────────────────────────────────────────────────────────────────────────────
def pagina_gestione_utenti(ws_utenti):
    st.markdown('<div class="main-title">👥 Gestione Utenti</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Aggiungi, disattiva e reimposta le password degli utenti</div>', unsafe_allow_html=True)

    df_u = carica_utenti(ws_utenti)

    # ── Lista utenti ──────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Utenti registrati</div>', unsafe_allow_html=True)
    if not df_u.empty:
        df_disp = df_u[["Username","Ruolo","Attivo","UltimoAccesso"]].copy()
        st.dataframe(df_disp, use_container_width=True, hide_index=True)
    else:
        st.info("Nessun utente trovato.")

    st.markdown("---")
    col_a, col_b = st.columns(2)

    # ── Reset password ────────────────────────────────────────────────────
    with col_a:
        st.markdown('<div class="section-header">🔑 Reimposta password</div>', unsafe_allow_html=True)
        if not df_u.empty:
            utenti_list = df_u[df_u["Username"] != ADMIN_USER]["Username"].tolist()
            u_reset = st.selectbox("Seleziona utente", utenti_list, key="u_reset")
            pwd_new  = st.text_input("Nuova password", type="password", key="pwd_new")
            pwd_new2 = st.text_input("Conferma",       type="password", key="pwd_new2")
            if st.button("🔑 Reimposta"):
                if not pwd_new: st.warning("Inserisci la nuova password.")
                elif pwd_new != pwd_new2: st.error("Le password non coincidono.")
                else:
                    if reset_password(ws_utenti, u_reset, pwd_new):
                        st.success(f"✅ Password di '{u_reset}' reimpostata.")
                    else: st.error("Utente non trovato.")
        else: st.info("Nessun utente da modificare.")

    # ── Attiva / disattiva ────────────────────────────────────────────────
    with col_b:
        st.markdown('<div class="section-header">🔒 Attiva / Disattiva account</div>', unsafe_allow_html=True)
        if not df_u.empty:
            utenti_list2 = df_u[df_u["Username"] != ADMIN_USER]["Username"].tolist()
            u_toggle = st.selectbox("Seleziona utente", utenti_list2, key="u_tog")
            stato_attuale = df_u[df_u["Username"]==u_toggle]["Attivo"].values[0] if len(df_u[df_u["Username"]==u_toggle])>0 else "SI"
            st.markdown(f"Stato attuale: **{'✅ Attivo' if stato_attuale=='SI' else '🔴 Disattivato'}**")
            c_on, c_off = st.columns(2)
            with c_on:
                if st.button("✅ Attiva"):
                    imposta_attivo(ws_utenti, u_toggle, True)
                    st.success(f"'{u_toggle}' attivato."); st.rerun()
            with c_off:
                if st.button("🔴 Disattiva"):
                    imposta_attivo(ws_utenti, u_toggle, False)
                    st.success(f"'{u_toggle}' disattivato."); st.rerun()
        else: st.info("Nessun utente da gestire.")

    # ── Aggiungi utente ───────────────────────────────────────────────────
    st.markdown('<div class="section-header">➕ Aggiungi utente</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1: nu_name = st.text_input("Username")
    with c2: nu_pwd  = st.text_input("Password", type="password", key="nu_pwd")
    with c3: nu_pwd2 = st.text_input("Conferma", type="password", key="nu_pwd2")
    with c4: nu_role = st.selectbox("Ruolo", ["partner","admin"])
    if st.button("➕ Crea utente"):
        if not nu_name.strip(): st.warning("Inserisci username.")
        elif not nu_pwd:        st.warning("Inserisci password.")
        elif nu_pwd != nu_pwd2: st.error("Le password non coincidono.")
        else:
            ok, msg = aggiungi_utente(ws_utenti, nu_name.strip(), nu_pwd, nu_role)
            if ok: st.success(f"✅ Utente '{nu_name}' creato con ruolo '{nu_role}'."); st.rerun()
            else:  st.error(msg)

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA CRUSCOTTO
# ─────────────────────────────────────────────────────────────────────────────
def pagina_cruscotto(ws_rend, ws_prev, ws_utenti):
    st.markdown('<div class="main-title">📊 Cruscotto Coordinatore</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Avanzamento rendicontazione vs preventivo</div>', unsafe_allow_html=True)

    df     = carica_df(ws_rend, INTESTAZIONE_REND)
    budget = carica_preventivo(ws_prev)

    if df.empty: st.info("Nessun dato ancora inserito."); return

    df["Importo (€)"] = pd.to_numeric(df["Importo (€)"], errors="coerce").fillna(0)
    df["Ore"]         = pd.to_numeric(df["Ore"],         errors="coerce").fillna(0)

    tot_fp = sum(v[0] for v in budget.values())
    tot_cp = sum(v[1] for v in budget.values())
    tot_fr = df[df["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
    tot_cr = df[df["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum()

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("💶 Fin. rendicontato",  f"€ {tot_fr:,.2f}", delta=f"{tot_fr/tot_fp*100:.1f}% prev." if tot_fp else "")
    c2.metric("🤝 Cofin. rendicontato",f"€ {tot_cr:,.2f}", delta=f"{tot_cr/tot_cp*100:.1f}% prev." if tot_cp else "")
    c3.metric("🕐 Ore totali",         f"{df['Ore'].sum():,.0f} h")
    c4.metric("📝 Voci inserite",      len(df))

    # Avanzamento per azione
    st.markdown('<div class="section-header">Avanzamento per Azione</div>', unsafe_allow_html=True)
    for azione,(pf,pc) in budget.items():
        df_az = df[df["Azione"]==azione]
        rf = df_az[df_az["Finanziamento/Cofinanziamento"]=="Finanziamento"]["Importo (€)"].sum()
        rc = df_az[df_az["Finanziamento/Cofinanziamento"]=="Cofinanziamento"]["Importo (€)"].sum()
        pf2 = rf/pf*100 if pf>0 else 0
        pc2 = rc/pc*100 if pc>0 else 0
        with st.expander(f"**{azione}**  —  Fin: {pf2:.1f}%  |  Cofin: {pc2:.1f}%"):
            ca,cb = st.columns(2)
            for col,label,rend,prev,perc,badge in [
                (ca,"FINANZIAMENTO",  rf,pf,pf2,"badge-fin"),
                (cb,"COFINANZIAMENTO",rc,pc,pc2,"badge-cofin")]:
                with col:
                    color = "#38a169" if perc<80 else ("#f59e0b" if perc<100 else "#e53e3e")
                    st.markdown(f"<span class='{badge}'>{label}</span>", unsafe_allow_html=True)
                    st.markdown(f"""<div style='margin:8px 0;'>
                      <div style='background:#e2e8f0;border-radius:999px;height:10px;'>
                        <div style='width:{min(perc,100):.1f}%;background:{color};height:100%;border-radius:999px;'></div>
                      </div><small style='color:#718096'>€ {rend:,.2f} / € {prev:,.2f}</small></div>""",
                        unsafe_allow_html=True)

    # Tabella riepilogativa partner
    st.markdown('<div class="section-header">Riepilogo per Partner</div>', unsafe_allow_html=True)
    df_sum = df.groupby(["Partner","Finanziamento/Cofinanziamento"])["Importo (€)"].sum().unstack(fill_value=0).reset_index()
    if "Finanziamento"   not in df_sum.columns: df_sum["Finanziamento"]   = 0
    if "Cofinanziamento" not in df_sum.columns: df_sum["Cofinanziamento"] = 0
    df_sum["Totale"] = df_sum["Finanziamento"] + df_sum["Cofinanziamento"]
    df_sum = df_sum.sort_values("Totale", ascending=False)
    grand_fin   = df_sum["Finanziamento"].sum()
    grand_cofin = df_sum["Cofinanziamento"].sum()
    grand_tot   = grand_fin + grand_cofin

    df_sum["% su Fin. totale"]   = df_sum["Finanziamento"].apply(  lambda x: f"{x/grand_fin*100:.1f}%"   if grand_fin>0   else "—")
    df_sum["% su Cofin. totale"] = df_sum["Cofinanziamento"].apply(lambda x: f"{x/grand_cofin*100:.1f}%" if grand_cofin>0 else "—")
    df_sum["% su Totale prog."]  = df_sum["Totale"].apply(         lambda x: f"{x/grand_tot*100:.1f}%"  if grand_tot>0  else "—")
    df_sum["Finanziamento"]   = df_sum["Finanziamento"].map(lambda x: f"€ {x:,.2f}")
    df_sum["Cofinanziamento"] = df_sum["Cofinanziamento"].map(lambda x: f"€ {x:,.2f}")
    df_sum["Totale"]          = df_sum["Totale"].map(lambda x: f"€ {x:,.2f}")

    st.dataframe(df_sum[["Partner","Finanziamento","% su Fin. totale",
                          "Cofinanziamento","% su Cofin. totale",
                          "Totale","% su Totale prog."]],
                 use_container_width=True, hide_index=True)

    # Dettaglio voci
    st.markdown('<div class="section-header">Dettaglio voci</div>', unsafe_allow_html=True)
    fp = st.selectbox("Partner", ["Tutti"]+df["Partner"].unique().tolist())
    fm = st.selectbox("Mese",    ["Tutti"]+MESI)
    df_v = df.copy()
    if fp!="Tutti": df_v = df_v[df_v["Partner"]==fp]
    if fm!="Tutti": df_v = df_v[df_v["Mese"]==fm]
    if not df_v.empty:
        df_v["Importo (€)"] = df_v["Importo (€)"].map(lambda x: f"€ {x:,.2f}")
        st.dataframe(df_v[["Timestamp","Partner","Mese","Azione","Tipo Costo","Descrizione",
                            "Ore","Costo Orario (€)","Importo (€)","Finanziamento/Cofinanziamento","Note"]],
                     use_container_width=True, hide_index=True)
        csv = df_v.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Esporta CSV", data=csv, file_name="rendicontazione.csv", mime="text/csv")
    else: st.info("Nessuna voce per i filtri selezionati.")

# ─────────────────────────────────────────────────────────────────────────────
# PAGINA EXPORT CONSUNTIVO
# ─────────────────────────────────────────────────────────────────────────────
def pagina_export_consuntivo(ws_rend, ws_prev):
    st.markdown('<div class="main-title">📤 Export Consuntivo</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Scarica il quadro logico a consuntivo in Excel o PDF</div>', unsafe_allow_html=True)

    df_rend = carica_df(ws_rend, INTESTAZIONE_REND)
    budget  = carica_preventivo(ws_prev)

    if df_rend.empty:
        st.markdown('<div class="alert-warn">⚠️ Nessun dato di rendicontazione ancora presente. Inserisci almeno qualche voce prima di generare il consuntivo.</div>', unsafe_allow_html=True)
        return

    df_rend["Importo (€)"] = pd.to_numeric(df_rend["Importo (€)"], errors="coerce").fillna(0)
    df_rend["Ore"]         = pd.to_numeric(df_rend["Ore"],         errors="coerce").fillna(0)

    # Anteprima riepilogo
    st.markdown('<div class="section-header">Anteprima riepilogo</div>', unsafe_allow_html=True)
    df_sum = df_rend.groupby(["Partner","Finanziamento/Cofinanziamento"])["Importo (€)"].sum().unstack(fill_value=0).reset_index()
    if "Finanziamento"   not in df_sum.columns: df_sum["Finanziamento"]   = 0
    if "Cofinanziamento" not in df_sum.columns: df_sum["Cofinanziamento"] = 0
    df_sum["Totale"] = df_sum["Finanziamento"] + df_sum["Cofinanziamento"]
    grand_tot = df_sum["Totale"].sum()
    df_sum["% Totale"] = df_sum["Totale"].apply(lambda x: f"{x/grand_tot*100:.1f}%" if grand_tot>0 else "—")
    df_sum_disp = df_sum.copy()
    for col in ["Finanziamento","Cofinanziamento","Totale"]:
        df_sum_disp[col] = df_sum_disp[col].map(lambda x: f"€ {x:,.2f}")
    st.dataframe(df_sum_disp, use_container_width=True, hide_index=True)
    st.markdown(f"**Totale generale rendicontato:** € {grand_tot:,.2f}")

    st.markdown("---")
    col_xl, col_pdf = st.columns(2)

    with col_xl:
        st.markdown("### 📊 Export Excel")
        st.markdown("3 fogli: *Consuntivo per Azione*, *Riepilogo Partner*, *Dati Grezzi*")
        if st.button("Genera Excel"):
            with st.spinner("Generazione Excel in corso..."):
                try:
                    xls_bytes = genera_excel_consuntivo(df_rend, budget, ANNO)
                    st.download_button(
                        label="⬇️ Scarica Excel",
                        data=xls_bytes,
                        file_name=f"consuntivo_{ANNO}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Errore generazione Excel: {e}")

    with col_pdf:
        st.markdown("### 📄 Export PDF")
        st.markdown("2 sezioni: *Riepilogo per Partner*, *Consuntivo per Azione* con scostamenti")
        if st.button("Genera PDF"):
            with st.spinner("Generazione PDF in corso..."):
                try:
                    pdf_bytes = genera_pdf_consuntivo(df_rend, budget, ANNO)
                    st.download_button(
                        label="⬇️ Scarica PDF",
                        data=pdf_bytes,
                        file_name=f"consuntivo_{ANNO}.pdf",
                        mime="application/pdf"
                    )
                except Exception as e:
                    st.error(f"Errore generazione PDF: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
client = get_gsheet_client()
if not client:
    st.error("Impossibile connettersi a Google Sheets. Verifica i secrets.")
    st.stop()

ws_rend, ws_costi, ws_prev, ws_utenti = get_worksheets(client)
if not ws_rend:
    st.error("Errore apertura fogli Google Sheets.")
    st.stop()

if not st.session_state.logged_in:
    mostra_login(ws_utenti)
else:
    pagina = mostra_sidebar()
    if   pagina == "📊 Cruscotto":           pagina_cruscotto(ws_rend, ws_prev, ws_utenti)
    elif pagina == "📋 Quadro logico":       pagina_quadro_logico(st.session_state.partner, st.session_state.ruolo)
    elif pagina == "⚙️ Gestione preventivo": pagina_gestione_preventivo(ws_costi, ws_prev, ws_utenti)
    elif pagina == "👥 Gestione utenti":     pagina_gestione_utenti(ws_utenti)
    elif pagina == "📤 Export consuntivo":   pagina_export_consuntivo(ws_rend, ws_prev)
    else:                                    pagina_inserimento(ws_rend, ws_costi, ws_prev, st.session_state.partner)
